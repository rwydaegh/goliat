"""Create Excel file matching CNR's Excel format for Thelonious and Eartha phantoms."""

import json
import logging
from pathlib import Path

import pandas as pd


def _auto_size_columns(worksheet, df: pd.DataFrame):
    """Auto-size column widths to fit content.

    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written to the worksheet
    """
    if df.empty:
        return

    # Calculate max width for each column
    for idx, column in enumerate(df.columns, start=1):
        # Get max length in column (including header)
        max_length = max(
            df[column].astype(str).map(len).max(),  # Max data length
            len(str(column)),  # Header length
        )
        # Add some padding and set minimum width
        adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, max 50
        worksheet.column_dimensions[worksheet.cell(1, idx).column_letter].width = adjusted_width


def _add_table_formatting(worksheet, df: pd.DataFrame, table_name: str):
    """Add Excel table formatting matching CNR's Excel format.

    Args:
        worksheet: openpyxl worksheet object
        df: DataFrame that was written to the worksheet
        table_name: Name for the table
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if df.empty:
        return

    # Calculate table range (A1 to last column + last row)
    num_rows = len(df) + 1  # +1 for header
    num_cols = len(df.columns)

    # Convert column number to letter (1 -> A, 2 -> B, etc.)
    def col_num_to_letter(n):
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result

    start_cell = "A1"
    end_col = col_num_to_letter(num_cols)
    end_cell = f"{end_col}{num_rows}"
    table_range = f"{start_cell}:{end_cell}"

    # Create table
    table = Table(displayName=table_name.replace(" ", "_"), ref=table_range)

    # Add default Excel green table style
    # TableStyleLight2 is the light green style, TableStyleMedium2 is medium green
    style = TableStyleInfo(
        name="TableStyleLight2",  # Light green table style (default Excel green)
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style

    # Add table to worksheet
    worksheet.add_table(table)


def get_target_power_mapping(config_path: str) -> dict[int, int]:
    """Read target power values from config file for each frequency."""
    with open(config_path, "r") as f:
        config = json.load(f)

    antenna_config = config.get("antenna_config", {})
    power_mapping = {}

    for freq_str, freq_config in antenna_config.items():
        if isinstance(freq_config, dict) and "target_power_mW" in freq_config:
            frequency_mhz = int(freq_str)
            power_mapping[frequency_mhz] = freq_config["target_power_mW"]

    return power_mapping


def map_placement_name(placement: str, sheet_type: str) -> str:
    """Map placement names to match CNR's Excel format.

    Args:
        placement: Original placement name from CSV
        sheet_type: One of 'fronteyes', 'belly', or 'cheek'

    Returns:
        Mapped placement name matching CNR's Excel format
    """
    if sheet_type == "fronteyes":
        # front_of_eyes_* placements stay the same
        if placement.startswith("front_of_eyes_"):
            return placement
    elif sheet_type == "belly":
        # by_belly_* -> belly_level_*
        if placement.startswith("by_belly_"):
            return placement.replace("by_belly_", "belly_level_")
    elif sheet_type == "cheek":
        # Map cheek placements to cheek_1, cheek_2, cheek_3, tilt_1, tilt_2, tilt_3
        mapping = {
            "by_cheek_tragus_cheek_base": "cheek_1",
            "by_cheek_tragus_cheek_up": "cheek_2",
            "by_cheek_tragus_cheek_down": "cheek_3",
            "by_cheek_tragus_tilt_base": "tilt_1",
            "by_cheek_tragus_tilt_up": "tilt_2",
            "by_cheek_tragus_tilt_down": "tilt_3",
        }
        return mapping.get(placement, placement)

    return placement


def create_sheet_data(
    df: pd.DataFrame,
    sheet_type: str,
    power_mapping: dict[int, int],
) -> pd.DataFrame:
    """Create data for one sheet matching CNR's Excel format.

    Args:
        df: Source DataFrame from CSV
        sheet_type: One of 'fronteyes', 'belly', or 'cheek'
        power_mapping: Dictionary mapping frequency (MHz) to target power (mW)

    Returns:
        DataFrame formatted to match CNR's Excel sheet format
    """
    # Filter placements based on sheet type
    if sheet_type == "fronteyes":
        filtered_df = df[df["placement"].str.startswith("front_of_eyes_")].copy()
    elif sheet_type == "belly":
        filtered_df = df[df["placement"].str.startswith("by_belly_")].copy()
    elif sheet_type == "cheek":
        filtered_df = df[df["placement"].str.startswith("by_cheek_tragus_")].copy()
    else:
        raise ValueError(f"Unknown sheet_type: {sheet_type}")

    if filtered_df.empty:
        return pd.DataFrame()

    # Map placement names to match CNR's Excel format
    filtered_df["placement"] = filtered_df["placement"].apply(lambda x: map_placement_name(x, sheet_type))

    # Create output DataFrame with column names matching CNR's Excel format
    result_df = pd.DataFrame()
    result_df["frequency_mhz"] = filtered_df["frequency_mhz"]
    result_df["placement"] = filtered_df["placement"]

    # Add Input Power column based on frequency (matching CNR's Excel format)
    result_df["Input Power (mW)"] = result_df["frequency_mhz"].map(power_mapping)

    # Map SAR columns to match CNR's Excel format
    result_df["SAR_wholebody (mW/kg)"] = filtered_df["SAR_whole_body"]
    result_df["SAR_head (mW/kg)"] = filtered_df["SAR_head"]
    result_df["SAR_trunk (mW/kg)"] = filtered_df["SAR_trunk"]
    result_df["psSAR10g_eyes (mW/kg)"] = filtered_df["psSAR10g_eyes"]
    result_df["psSAR10g_skin (mW/kg)"] = filtered_df["psSAR10g_skin"]
    result_df["psSAR10g_brain (mW/kg)"] = filtered_df["psSAR10g_brain"]

    # Sort by frequency and placement
    result_df = result_df.sort_values(by=["frequency_mhz", "placement"]).reset_index(drop=True)

    return result_df


def create_cnr_excel(
    phantom_name: str,
    csv_path: str,
    config_path: str,
    output_path: str,
):
    """Create Excel file matching CNR's Excel format for a phantom.

    Args:
        phantom_name: Name of the phantom (e.g., 'thelonious', 'eartha')
        csv_path: Path to normalized_results_detailed.csv
        config_path: Path to config JSON file with antenna_config
        output_path: Path where Excel file should be saved
    """
    # Read CSV data
    df = pd.read_csv(csv_path)

    # Get power mapping from config
    power_mapping = get_target_power_mapping(config_path)

    # Create data for each sheet
    sheet_fronteyes = create_sheet_data(df, "fronteyes", power_mapping)
    sheet_belly = create_sheet_data(df, "belly", power_mapping)
    sheet_cheek = create_sheet_data(df, "cheek", power_mapping)

    # Create Excel file with multiple sheets
    sheet_name = f"{phantom_name.capitalize()}_fronteyes"
    belly_name = f"{phantom_name.capitalize()}_belly"
    cheek_name = f"{phantom_name.capitalize()}_cheek"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Write sheets in order: fronteyes (sheet 1), belly (sheet 2), cheek (sheet 3)
        sheet_fronteyes.to_excel(writer, sheet_name=sheet_name, index=False)
        sheet_belly.to_excel(writer, sheet_name=belly_name, index=False)
        sheet_cheek.to_excel(writer, sheet_name=cheek_name, index=False)

    print(f"Created Excel file: {output_path}")
    print(f"  - Sheet 1 ({sheet_name}): {len(sheet_fronteyes)} rows")
    print(f"  - Sheet 2 ({belly_name}): {len(sheet_belly)} rows")
    print(f"  - Sheet 3 ({cheek_name}): {len(sheet_cheek)} rows")


def main():
    """Main function to create a single Excel file matching CNR's Excel format for both phantoms."""
    # Paths
    # __file__ is in goliat/goliat/analysis/, so we need to go up to goliat/
    base_dir = Path(__file__).parent.parent.parent
    results_dir = base_dir / "results" / "near_field"
    config_path = base_dir / "configs" / "near_field_config.json"

    phantoms = ["thelonious", "eartha"]

    # Create a single Excel file with all sheets
    output_path = results_dir / "Final_Data_UGent.xlsx"

    # Get power mapping from config
    power_mapping = get_target_power_mapping(str(config_path))

    # Create Excel writer
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for phantom_name in phantoms:
            csv_path = results_dir / phantom_name / "normalized_results_detailed.csv"

            if not csv_path.exists():
                logging.getLogger("progress").warning(f"CSV file not found: {csv_path}", extra={"log_type": "warning"})
                continue

            logging.getLogger("progress").info(f"  Processing: {phantom_name.capitalize()}...", extra={"log_type": "progress"})

            # Read CSV data
            df = pd.read_csv(csv_path)

            # Create data for each sheet
            sheet_fronteyes = create_sheet_data(df, "fronteyes", power_mapping)
            sheet_belly = create_sheet_data(df, "belly", power_mapping)
            sheet_cheek = create_sheet_data(df, "cheek", power_mapping)

            # Create sheet names
            sheet_name_fronteyes = f"{phantom_name.capitalize()}_fronteyes"
            sheet_name_belly = f"{phantom_name.capitalize()}_belly"
            sheet_name_cheek = f"{phantom_name.capitalize()}_cheek"

            # Write sheets
            sheet_fronteyes.to_excel(writer, sheet_name=sheet_name_fronteyes, index=False)
            sheet_belly.to_excel(writer, sheet_name=sheet_name_belly, index=False)
            sheet_cheek.to_excel(writer, sheet_name=sheet_name_cheek, index=False)

            # Add table formatting to each sheet
            workbook = writer.book

            # Format fronteyes sheet as table
            if not sheet_fronteyes.empty:
                ws_fronteyes = workbook[sheet_name_fronteyes]
                _add_table_formatting(ws_fronteyes, sheet_fronteyes, sheet_name_fronteyes)
                _auto_size_columns(ws_fronteyes, sheet_fronteyes)

            # Format belly sheet as table
            if not sheet_belly.empty:
                ws_belly = workbook[sheet_name_belly]
                _add_table_formatting(ws_belly, sheet_belly, sheet_name_belly)
                _auto_size_columns(ws_belly, sheet_belly)

            # Format cheek sheet as table
            if not sheet_cheek.empty:
                ws_cheek = workbook[sheet_name_cheek]
                _add_table_formatting(ws_cheek, sheet_cheek, sheet_name_cheek)
                _auto_size_columns(ws_cheek, sheet_cheek)

            logging.getLogger("progress").info(f"    - {sheet_name_fronteyes}: {len(sheet_fronteyes)} rows", extra={"log_type": "verbose"})
            logging.getLogger("progress").info(f"    - {sheet_name_belly}: {len(sheet_belly)} rows", extra={"log_type": "verbose"})
            logging.getLogger("progress").info(f"    - {sheet_name_cheek}: {len(sheet_cheek)} rows", extra={"log_type": "verbose"})

    logging.getLogger("progress").info(f"  Created: {output_path.name}", extra={"log_type": "success"})


if __name__ == "__main__":
    main()
